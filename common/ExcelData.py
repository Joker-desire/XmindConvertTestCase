# -*- coding utf-8 -*-
# @Time    : 2020/11/14 10:17
# @Author  : DesireYang
# @Email   : yangyin1106@163.com
# @File    : ExcelData.py
# Software : PyCharm
# Explain  : 写入到Excel
import os

import openpyxl


class ExcelData:

    def __init__(self, file_name=None, sheetName="Sheet"):
        """
        初始化方法
        :param file_name:
        :param sheetName:
        """
        self.fileName = file_name
        self.sheetName = sheetName

    def create_excel_and_set_title(self, titles: list = None):
        """
        创建Excel文件，如果存在就先删除后创建（保证Excel数据无异常）
        设置Excel title
        :param titles: Excel表单头标题
        """
        # 判断是否存在（如果存在就删除后重新创建，保证Excel数据无异常）
        if os.path.exists(self.fileName):
            # 删除文件
            os.remove(self.fileName)
        # 创建文件
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = self.sheetName
        wb.save(self.fileName)
        if titles is not None:
            for i in enumerate(titles):
                self.write_excel_data(1, int(i[0]) + 1, value=i[1])

    def __open_excel(self):
        """
        打开工作簿，选中表单
        """
        self.wb = openpyxl.load_workbook(self.fileName)
        self.sh = self.wb[self.sheetName]

    def __save(self):
        """
        保存工作簿对象的方法
        """
        self.wb.save(self.fileName)
        self.wb.close()  # 释放内存

    def write_excel_data(self, row, column, value):
        """
        写入数据
        :param row: 行
        :param column: 列
        :param value: 数据
        """
        self.__open_excel()
        # 指定行列进行写入数据
        self.sh.cell(row=row, column=column, value=value)

        # 保存
        self.__save()


if __name__ == '__main__':
    fileName = "../s.xlsx"
    we = ExcelData(fileName)
    we.create_excel_and_set_title()
